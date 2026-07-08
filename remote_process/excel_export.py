from config_path import PATH_HARVEST
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# df = ... # ton dataframe à importer (déjà préparé, issu d'un merge avec indicator=True
#            par ex., donc avec les colonnes geo_admin_new et _merge)

def update_geo_admin(file_path: str, sheet, df: pd.DataFrame):
    wb = load_workbook(file_path, keep_vba=True)

    backup_name = sheet+"_"+datetime.now().strftime("%Y_%m")  # ex: 2026_06, calculé automatiquement

    # 1. Sauvegarder l'onglet existant sous un autre nom (copie avant écrasement)
    if sheet in wb.sheetnames:
        if backup_name in wb.sheetnames:
            del wb[backup_name]  # évite un conflit si on relance le script
        src_ws = wb[sheet]
        backup_ws = wb.copy_worksheet(src_ws)  # copie valeurs + la plupart des styles
        backup_ws.title = backup_name


    # 3. Écraser l'onglet geo_admin (on le supprime et on le recrée au même endroit)
    if sheet in wb.sheetnames:
        idx = wb.sheetnames.index(sheet)
        del wb[sheet]
    
    ws = wb.create_sheet(sheet)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(file_path)